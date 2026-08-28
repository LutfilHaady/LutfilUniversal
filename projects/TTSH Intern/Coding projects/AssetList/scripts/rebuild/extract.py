import openpyxl
from mapping import SHEETS, ALL_FIELDS, BLOCKS


def read_raw_rows(source_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(source_path, data_only=True)
    rows = []
    for sheet in SHEETS:
        ws = wb[sheet["name"]]
        for r in range(sheet["first"], sheet["last"] + 1):
            row = {"dept_code": sheet["code"], "source_row": r}
            for field in ALL_FIELDS:
                col = sheet["cols"].get(field)
                row[field] = ws[f"{col}{r}"].value if col else None
            rows.append(row)
    return rows


def clean_text(value):
    if value is None:
        return None
    text = str(value).replace("\xa0", "").strip()
    return text if text else None


def classify_category(asset_type):
    text = (asset_type or "").lower()
    has = lambda kw: kw in text
    if has("laptop"):
        return "Laptop"
    if has("tablet") or has("ipad"):
        return "Tablet"
    if has("monitor") and has("cpu"):
        return "Monitor + CPU"
    if has("cpu") or has("desktop"):
        return "CPU"
    if has("printer"):
        return "Printer"
    if has("scanner"):
        return "Document Scanner"
    if has("topaz") or has("signature pad") or has("singnature pad") or has("sign pad"):
        return "Topaz Signature Pad"
    if has("nets"):
        return "Nets Machine"
    return "Other (Review)"


def _block_for(dept_code, source_row):
    for first, last, label, is_junk in BLOCKS.get(dept_code, []):
        if first <= source_row <= last:
            return label, is_junk
    return "", False


def normalize_rows(raw_rows):
    asset_rows, junk_rows = [], []
    counters = {}
    for row in raw_rows:
        # _block_for already encodes junk status from the verified row ranges in
        # mapping.BLOCKS (PSOINF 37-43, AONCID 53-54); depts with no BLOCKS entry
        # always return is_junk=False.
        label, is_junk = _block_for(row["dept_code"], row["source_row"])
        if is_junk:
            junk_rows.append(row)
            continue
        row = dict(row)
        row["serial"] = clean_text(row["serial"])
        row["itd_tag"] = clean_text(row["itd_tag"])
        row["asset_group"] = label
        row["asset_category"] = classify_category(row["asset_type"])
        n = counters.get(row["dept_code"], 0) + 1
        counters[row["dept_code"]] = n
        row["asset_id"] = f"{row['dept_code']}-{n:03d}"
        asset_rows.append(row)
    return asset_rows, junk_rows


def find_duplicate_serials(asset_rows):
    by_serial = {}
    for row in asset_rows:
        if not row["serial"]:
            continue
        by_serial.setdefault(row["serial"], []).append(f"{row['dept_code']} r{row['source_row']}")
    return [(serial, locs) for serial, locs in by_serial.items() if len(locs) > 1]


def read_plain_comments(source_path):
    wb = openpyxl.load_workbook(source_path)
    ws = wb["EDFC (Guo wei)"]
    out = []
    for cell_ref, source_row, field in [("G4", 4, "host"), ("G5", 5, "host"),
                                         ("D12", 12, "itd_tag"), ("D13", 13, "itd_tag")]:
        c = ws[cell_ref]
        if c.comment:
            out.append(("EDFC", source_row, field, c.comment.author or "Unknown", c.comment.text.strip()))
    return out


if __name__ == "__main__":
    rows = read_raw_rows("../../July 2026 FC IT Assets Management list (Cost centre checked).xlsx")
    assert len(rows) == 171, f"expected 171 raw rows, got {len(rows)}"
    by_dept = {}
    for row in rows:
        by_dept[row["dept_code"]] = by_dept.get(row["dept_code"], 0) + 1
    expected = {"A1L5": 28, "ICHPSO": 27, "PSOINF": 41, "AONCID": 52, "EDFC": 23}
    assert by_dept == expected, f"expected {expected}, got {by_dept}"
    print("OK: 171 raw rows,", by_dept)

    asset_rows, junk_rows = normalize_rows(rows)
    assert len(junk_rows) == 9, f"expected 9 junk rows, got {len(junk_rows)}"
    assert len(asset_rows) == 162, f"expected 162 asset rows, got {len(asset_rows)}"
    ids = [r["asset_id"] for r in asset_rows]
    assert len(ids) == len(set(ids)), "Asset IDs are not unique"
    cats = {}
    for r in asset_rows:
        cats[r["asset_category"]] = cats.get(r["asset_category"], 0) + 1
    print("OK: 162 asset rows, 9 junk rows, IDs unique")
    print("Category counts:", cats)

    dups = find_duplicate_serials(asset_rows)
    print("Duplicate serials:", len(dups))
    for serial, locs in dups:
        print("  ", serial, locs)

    plain_comments = read_plain_comments(
        "../../July 2026 FC IT Assets Management list (Cost centre checked).xlsx")
    print("Plain comments found on EDFC:", len(plain_comments))
    for c in plain_comments:
        print("  ", c)
